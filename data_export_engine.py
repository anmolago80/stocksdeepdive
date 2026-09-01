"""
data_export_engine.py

Services batch 2, Part 3 (2026-09-01): "Download data" - CSV/Excel export
of numbers already on screen. Pure re-shaping of whatever other engines
already computed (deep_dive_engine.analyze(), auto_compounder_engine.
build_sections(), reverse_dcf_engine.compute(), peer_context.compute(),
and the plain pandas tables behind the Scanner/Comparison/My Portfolio
pages) - no new network call, no new calculation, and none of those
engines are touched or re-implemented here. No Anthropic API call
anywhere in this feature.

Two Deep Dive export shapes, both built entirely in memory (no temp
files, no disk write):

  build_deep_dive_workbook(...) -> bytes (.xlsx via openpyxl)
      One sheet per Compounder View tab (Fundamentals, Value vs Book,
      Retained Earnings, Earnings Trends, Cost of Capital, Fair Value),
      each with the same label / value / comment / flagged content
      compounder_ui.render_section() puts on screen for that tab (see
      _metric_rows below - it reads the exact same section["metrics"]
      list render_section() itself iterates, so the two can never drift
      apart), plus "Measured over" horizon windows on the Retained
      Earnings sheet (same source as that tab's own caption - see
      _retained_earnings_windows). Also a "Valuation" sheet (price,
      intrinsic value, MOS, the DCF inputs used, and Part 1's reverse-DCF
      numbers when available), a "Scores" sheet (the five scores plus
      each one's contribution to the Long Score), and a "Source" sheet
      (as-of date, data-provider note, site attribution/disclaimer - the
      same ATTRIBUTION/PLAIN_DISCLAIMER text every other public surface
      already uses, see snapshot_render.py).

  build_deep_dive_csv(...) -> bytes (.csv)
      The Valuation + Scores sheets above, flattened to one Sheet/Label/
      Value table - the CSV variant the spec calls for alongside the
      full workbook.

Paywall: `subscriber` (bool) is resolved by the CALLER from
paywall_engine (PAYWALL_ENABLED off, or is_logged_in()+is_subscribed())
- this module has no Streamlit dependency and never touches paywall_engine
itself, so it can't render a gate; it only respects the flag it's given.
A non-subscriber still gets a "Fair Value" sheet (so the workbook's tab
order matches the on-screen tab order exactly) but with a one-line note
instead of the real numbers, mirroring the exact "Fair Value" omission
_dd_ask_context() already applies before handing acv_sections to the AI
Ask box (see app.py) - this is not a second, easier way to see paywalled
numbers than opening the tab.

table_to_csv_bytes(df) is a tiny shared helper for the other three
export buttons (Scanner/Comparison/My Portfolio), which are already
plain pandas DataFrames on screen and need no per-tab reshaping - kept
here purely so every "Download ... (CSV)" button on the site goes
through one small, tested function instead of four separate ad hoc
`df.to_csv()` call sites.
"""

import io
import re
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

import snapshot_render

SECTION_ORDER = [
    "Fundamentals", "Value vs Book", "Retained Earnings",
    "Earnings Trends", "Cost of Capital", "Fair Value",
]

_HEADER_FONT = Font(bold=True)
_WRAP_TOP = Alignment(wrap_text=True, vertical="top")
_TITLE_FONT = Font(bold=True, size=13)


def _vc_horizon_sort_key(h):
    """Same span-order as compounder_ui._vc_horizon_sort_key (numeric
    span ascending, TTM last) - reimplemented rather than imported
    across modules purely to keep this export module import-light (no
    Streamlit dependency at all); the two must stay logically identical
    since both are describing the same value_created dict."""
    if h.startswith("TTM"):
        return (999, h)
    m = re.match(r"(\d+)Y", h)
    return (int(m.group(1)) if m else 500, h)


def _autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_rows(ws, headers, rows, widths, wrap_cols=()):
    """headers: list[str]. rows: list[list]. widths: list[int] matching
    headers. wrap_cols: 0-based column indices that should wrap text and
    align to the top (used for long comment/note text)."""
    ws.append(headers)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
    for row in rows:
        ws.append(row)
    for r in range(2, len(rows) + 2):
        for c in wrap_cols:
            ws.cell(row=r, column=c + 1).alignment = _WRAP_TOP
    _autosize(ws, widths)
    ws.freeze_panes = "A2"


def _metric_rows(section, ticker):
    """[(label, value, comment, flagged), ...] for one build_sections()
    section dict, in the section's own metrics-list order - the exact
    same list compounder_ui.render_section() iterates (colour-coded and
    plain metrics together, un-split here since a spreadsheet row has no
    gauge-vs-metric visual distinction to preserve)."""
    rows = []
    for m in (section or {}).get("metrics", []):
        values = m.get("values") or {}
        value = values.get(ticker)
        if value is None and len(values) == 1:
            # Tolerate a ticker-casing mismatch (build_sections() always
            # keys by the exact ticker it was called with, so this is a
            # defensive fallback, not the expected path).
            value = next(iter(values.values()))
        rows.append((m.get("label"), value, m.get("comment") or "", bool(m.get("flagged"))))
    return rows


def _retained_earnings_windows(section, ticker):
    """['1Y = Jun 2024 -> Jun 2025', ...] - the same "Measured over" text
    render_section() builds from value_created's per-horizon "window"
    strings (see compounder_ui.render_section's own _vc_windows). Hand-
    built (non-live) sections carry no window strings, so this returns
    [] for those, same as the on-screen caption simply not appearing."""
    entry = ((section or {}).get("value_created") or {}).get(ticker) or {}
    windows = []
    for h in sorted([k for k in entry if not k.startswith("_")], key=_vc_horizon_sort_key):
        v = entry.get(h)
        if isinstance(v, dict) and v.get("window"):
            windows.append(f"{h} = {v['window']}")
    return windows


def _fmt_value(value):
    """Cell-friendly value: real numbers/bools/strings pass through
    unchanged (so Excel sees an actual number, not text - the workbook's
    own "opens with correct types" requirement), None becomes a blank
    cell rather than the text "None"."""
    if value is None:
        return None
    return value


# -----------------------------------
# Valuation / Scores sheets - built from the Deep Dive's own `dd` dict
# (deep_dive_engine.analyze()'s return shape) plus, when available,
# reverse_dcf_engine.compute()'s result and peer_context.compute()'s
# result - the exact three sources the Deep Dive page itself renders
# (Margin of Safety block, the reverse-DCF card, the Peer context
# block). None of the three is recomputed here; the caller passes in
# whatever it already has on hand for this page view.
# -----------------------------------

def _valuation_rows(dd, reverse_dcf=None):
    rows = [
        ("Price", _fmt_value(dd.get("price"))),
        ("Currency", _fmt_value(dd.get("currency"))),
        ("Valuation", _fmt_value(dd.get("valuation"))),
        ("Intrinsic Value (DCF)", _fmt_value(dd.get("intrinsic_value"))),
        ("Margin of Safety %", _fmt_value(dd.get("mos"))),
        ("Intrinsic value rests on a default/estimate", bool(dd.get("value_default"))),
        ("DCF growth rate used %", _fmt_value(dd.get("dcf_growth"))),
        ("DCF discount rate used %", _fmt_value(dd.get("dcf_discount"))),
        ("DCF perpetual rate used %", _fmt_value(dd.get("dcf_perpetual"))),
        ("Growth governor", _fmt_value(dd.get("growth_governor"))),
    ]
    if reverse_dcf and reverse_dcf.get("ok"):
        implied = reverse_dcf.get("implied_growth")
        model = reverse_dcf.get("model_growth")
        rows.extend([
            ("Implied growth (reverse DCF) %",
             round(implied * 100, 1) if implied is not None else None),
            ("Model growth assumed (reverse DCF) %",
             round(model * 100, 1) if model is not None else None),
            ("Reverse DCF hit the search boundary",
             reverse_dcf.get("implied_growth_capped") or ""),
        ])
    return rows


def _scores_rows(dd):
    rows = [
        ("Value Score", _fmt_value(dd.get("long_score"))),
        ("Quality", _fmt_value(dd.get("quality_score"))),
        ("Quality label", _fmt_value(dd.get("quality_label"))),
        ("Quality rests on a default/estimate", bool(dd.get("quality_default"))),
        ("Psychology", _fmt_value(dd.get("psychology"))),
        ("Psychology sentiment", _fmt_value(dd.get("psychology_sentiment"))),
        ("Discovery (attention)", _fmt_value(dd.get("discovery"))),
        ("Discovery label", _fmt_value(dd.get("discovery_label"))),
        ("Moat", _fmt_value(dd.get("moat"))),
        ("Moat band", _fmt_value(dd.get("moat_band_label"))),
        ("Moat erosion", _fmt_value(dd.get("moat_erosion"))),
    ]
    contributions = dd.get("contributions") or {}
    for label, value in contributions.items():
        rows.append((f"Value Score driver - {label}", _fmt_value(value)))
    return rows


def _write_valuation_sheet(wb, dd, reverse_dcf):
    ws = wb.create_sheet("Valuation")
    _write_rows(
        ws, ["Metric", "Value"],
        [[label, value] for label, value in _valuation_rows(dd, reverse_dcf)],
        widths=[34, 20],
    )


def _write_scores_sheet(wb, dd):
    ws = wb.create_sheet("Scores")
    _write_rows(
        ws, ["Metric", "Value"],
        [[label, value] for label, value in _scores_rows(dd)],
        widths=[34, 20],
    )


def _write_section_sheet(wb, section_label, section, ticker, locked_note=None):
    ws = wb.create_sheet(section_label[:31])
    if locked_note:
        ws.append([locked_note])
        ws["A1"].font = _TITLE_FONT
        ws["A1"].alignment = _WRAP_TOP
        _autosize(ws, [80])
        return
    rows = _metric_rows(section, ticker)
    _write_rows(
        ws, ["Label", "Value", "Flagged (estimate)", "Comment"],
        [[label, _fmt_value(value), "Yes" if flagged else "", comment]
         for label, value, comment, flagged in rows],
        widths=[34, 16, 16, 80], wrap_cols=(3,),
    )
    if section_label == "Retained Earnings":
        windows = _retained_earnings_windows(section, ticker)
        if windows:
            start = ws.max_row + 2
            ws.cell(row=start, column=1, value="Measured over").font = _HEADER_FONT
            for i, w in enumerate(windows):
                ws.cell(row=start + 1 + i, column=1, value=w)


def _write_source_sheet(wb, ticker, subscriber, generated_note=None):
    ws = wb.create_sheet("Source")
    rows = [
        ["Ticker", ticker],
        ["Generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")],
        ["Data provider", "Reported company financials and market data via the site's usual data provider - see /methodology."],
        ["Framing", "Descriptions of calculations, not recommendations."],
        ["Attribution", snapshot_render.ATTRIBUTION],
        ["Disclaimer", snapshot_render.PLAIN_DISCLAIMER],
    ]
    if not subscriber:
        rows.append([
            "Fair Value sheet",
            "Omitted - Fair Value is a paywalled section; subscribe on the site to include it in this workbook.",
        ])
    if generated_note:
        rows.append(["Note", generated_note])
    _write_rows(ws, ["Field", "Value"], rows, widths=[20, 90], wrap_cols=(1,))


def deep_dive_filename(ticker, ext="xlsx"):
    date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    return f"StocksDeepDive_{ticker}_{date}.{ext}"


def build_deep_dive_workbook(ticker, dd, acv_sections=None, reverse_dcf=None, subscriber=True):
    """In-memory .xlsx bytes for one ticker's Deep Dive - see module
    docstring for the sheet list. `acv_sections` is
    auto_compounder_engine.build_sections()'s return value (or None/{}
    if it couldn't be computed, e.g. Compounder View itself failed for
    this ticker on this view - the six tab sheets are then written with
    a plain "no data" placeholder instead of being skipped, so the
    workbook's tab order is always the six tabs in the same order every
    time). `reverse_dcf` is reverse_dcf_engine.compute()'s result (or
    None if Part 1's card doesn't apply to this ticker, e.g. a P/E-blend
    name with no positive FCF base)."""
    wb = Workbook()
    wb.remove(wb.active)

    _write_valuation_sheet(wb, dd, reverse_dcf)
    _write_scores_sheet(wb, dd)

    acv_sections = acv_sections or {}
    for label in SECTION_ORDER:
        section = acv_sections.get(label)
        if label == "Fair Value" and not subscriber:
            _write_section_sheet(
                wb, label, None, ticker,
                locked_note=(
                    "Fair Value is a paywalled section on this site - "
                    "subscribe to include the full valuation-methods "
                    "breakdown in this workbook."
                ),
            )
            continue
        if not section:
            _write_section_sheet(
                wb, label, None, ticker,
                locked_note=f"No {label} data available for {ticker} on this view.",
            )
            continue
        _write_section_sheet(wb, label, section, ticker)

    _write_source_sheet(wb, ticker, subscriber)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_deep_dive_csv(ticker, dd, reverse_dcf=None):
    """In-memory .csv bytes: the Valuation + Scores sheets above,
    flattened to one Sheet/Label/Value table - the lightweight companion
    to build_deep_dive_workbook() the spec calls for alongside the full
    workbook."""
    import csv

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Sheet", "Label", "Value"])
    for label, value in _valuation_rows(dd, reverse_dcf):
        writer.writerow(["Valuation", label, "" if value is None else value])
    for label, value in _scores_rows(dd):
        writer.writerow(["Scores", label, "" if value is None else value])
    return buf.getvalue().encode("utf-8")


def table_to_csv_bytes(df):
    """Shared helper for the Scanner/Comparison/My Portfolio "Download...
    (CSV)" buttons - a plain pandas DataFrame -> in-memory CSV bytes, no
    reshaping (those tables are already exactly what's on screen)."""
    return df.to_csv(index=False).encode("utf-8")
