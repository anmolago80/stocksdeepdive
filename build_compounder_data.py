"""
Builds compounder_data.json from the source research workbook.

This can be run either way: as a standalone offline script (see the
command below), or live from the app's admin-only refresh panel, which
calls build() directly against an uploaded workbook. If a Railway Volume
is attached to the service (see _cp_data_dir() below), the rebuilt files
are written straight to it and survive redeploys/restarts on their own --
no further action needed. Without a volume, Railway's filesystem is
ephemeral, so whoever triggers a rebuild needs to download the resulting
compounder_data.json and commit/redeploy it for the update to survive the
app's next restart - see the admin panel's own docstring in app.py for
that flow.

    python3 build_compounder_data.py [path-to-xlsx]

Only the seven "Stock name in the watchlist" sheets between Stock Analysis
and Company Potential are read. For every metric column in those sheets
that Andrew himself attached an explanatory comment to in Excel, this
pulls: the header text, his comment, the cached value for every ticker
that has real data, and - where his comment actually spells out red/amber/
green style interpretation bands - a curated set of colour thresholds so
the app can colour-code the same way he does by eye in the spreadsheet.

Columns whose comment is purely explanatory (no numeric bands) are still
included so the app can show them as plain (uncoloured) numbers - nothing
gets invented that Andrew didn't already write down himself.
"""

import hashlib
import json
import os
import sys
from datetime import datetime, timezone

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

DEFAULT_PATH = (
    "/mnt/user-data/uploads/Blogs and Dashboard Investment/"
    "Shares - Invested and Investigation - Rev 3 - SMSF.xlsx"
)
OUT_PATH = "compounder_data.json"


def _cp_data_dir():
    """Directory where the two persisted Rational Compounder files
    (compounder_data.json, company_potential_corrections.json) actually
    live at runtime. Prefers a Railway Volume if one is attached to this
    service -- Railway auto-sets RAILWAY_VOLUME_MOUNT_PATH the moment a
    volume is mounted, no code-side coordination needed -- so admin-panel
    rebuilds persist across redeploys/restarts on their own. Falls back to
    this file's own directory (the git-tracked copies) for local runs, the
    CLI usage below, or before a volume is attached."""
    return os.environ.get("RAILWAY_VOLUME_MOUNT_PATH") or os.path.dirname(__file__)


def _cp_seed_from_repo_if_missing(filename):
    """The very first time a volume is attached, it's empty -- copy the
    git-tracked copy of `filename` (living next to this script) over as a
    starting point so the site doesn't show blank data (or lose the
    corrections cache) until the next rebuild. No-op once the volume has
    its own copy, and a no-op entirely when there's no volume (data dir ==
    script dir already)."""
    data_dir = _cp_data_dir()
    repo_path = os.path.join(os.path.dirname(__file__), filename)
    volume_path = os.path.join(data_dir, filename)
    if volume_path == repo_path:
        return
    if os.path.exists(volume_path) or not os.path.exists(repo_path):
        return
    try:
        with open(repo_path, "rb") as src, open(volume_path, "wb") as dst:
            dst.write(src.read())
    except OSError:
        pass

# Ticker column per sheet, and which rows currently hold real (non-error)
# stock data. Rows are 4, 5, 6, ... - extend this list as Andrew finishes
# researching more names; anything past the workbook's real rows is just
# skipped (formula errors / blanks) rather than erroring out.
TICKER_COL = {
    "Stock Analysis": 4,
    "IV2BV": 4,
    "Dividend Ratio": 4,
    "Earnings Analysis": 4,
    "Cost of Capital Analysis": 4,
    "Valuation": 5,
    "Company Potential": 2,
}
MAX_DATA_ROWS = 200  # scan this many rows per sheet looking for real tickers
HEADER_ROWS = {
    "Stock Analysis": [3],
    "IV2BV": [3],
    "Dividend Ratio": [3],
    "Earnings Analysis": [3],
    "Cost of Capital Analysis": [1, 3],
    "Valuation": [3],
    "Company Potential": [1, 2, 3],
}

# Section order + display names (dropdown 2), decided with Andrew - raw
# sheet names kept only as the internal lookup key.
SECTIONS = [
    ("Fundamentals", "Stock Analysis"),
    ("Value vs Book", "IV2BV"),
    ("Retained Earnings", "Dividend Ratio"),
    ("Earnings Trends", "Earnings Analysis"),
    ("Cost of Capital", "Cost of Capital Analysis"),
    ("Fair Value", "Valuation"),
    ("Company Potential", "Company Potential"),
]

# format: how to display the raw number
#   pct  -> value * 100, one decimal, "%"
#   x    -> two decimals, "x" suffix
#   num  -> plain number, comma thousands
#   cur  -> "$" + comma thousands (0 decimals if large, else 2)
#
# thresholds: list of (lo, hi, color, band_label) using the RAW value (same
# units as the sheet, e.g. 0.06 not 6). lo/hi = None means open-ended.
# color in {"red", "amber", "green", "blue"} - blue = the extra "very
# high / special" 4th tier a few of Andrew's comments use (Market Cap/NTA,
# Interest Coverage, Earning Yield, EV/FCF).
#
# Every band below is taken directly from the header comment Andrew wrote
# on that column - nothing here is an invented cutoff.
METRIC_META = {
    ("Stock Analysis", "O"): dict(format="pct", thresholds=[
        (None, 0.04, "red", "Expensive"),
        (0.04, 0.06, "amber", "Fair value"),
        (0.06, 0.08, "green", "Reasonably attractive"),
        (0.08, None, "blue", "Cheap (or risky)"),
    ]),
    ("Stock Analysis", "P"): dict(format="x", thresholds=[
        (None, 1, "amber", "Very cheap (or troubled)"),
        (1, 3, "green", "Reasonable / fair"),
        (3, 6, "amber", "Growth premium"),
        (6, None, "red", "Expensive"),
    ]),
    ("Stock Analysis", "R"): dict(format="cur"),
    ("Stock Analysis", "U"): dict(format="cur"),
    ("Stock Analysis", "V"): dict(format="cur"),
    ("Stock Analysis", "Z"): dict(format="cur"),
    ("Stock Analysis", "AA"): dict(format="x", thresholds=[
        (0.5, 1.5, "red", "Near/below tangible backing"),
        (1.5, 3.0, "amber", "Fairly valued, stable"),
        (3.0, 6.0, "green", "Premium, high quality"),
        (6.0, 15.0, "blue", "Very high (IP-driven / growth)"),
    ]),
    ("Stock Analysis", "AG"): dict(format="cur"),
    ("Stock Analysis", "AH"): dict(format="cur"),
    ("Stock Analysis", "AI"): dict(format="x", thresholds=[
        (None, 1.0, "red", "Extreme risk"),
        (1.0, 1.5, "red", "High risk"),
        (1.5, 3.0, "amber", "Thin buffer"),
        (3.0, 5.0, "green", "Comfortable coverage"),
        (5.0, None, "blue", "Very strong buffer"),
    ]),
    ("Stock Analysis", "AS"): dict(format="x", thresholds=[
        (0.5, None, "green", "Strong ability to cover debt"),
        (0.2, 0.5, "amber", "Moderate coverage"),
        (None, 0.2, "red", "Weak, higher risk"),
    ]),
    ("Stock Analysis", "AT"): dict(format="x", thresholds=[
        (None, 10, "blue", "Very cheap"),
        (10, 15, "green", "Attractive / good value"),
        (15, 20, "amber", "Fair / market level"),
        (20, None, "red", "Expensive"),
    ]),
    ("Stock Analysis", "AX"): dict(format="pct"),
    ("Stock Analysis", "AY"): dict(format="pct", thresholds=[
        (None, 0.03, "red", "Expensive, weak cash return"),
        (0.03, 0.06, "amber", "Fair value"),
        (0.06, None, "green", "Attractive, strong cash return"),
    ]),
    ("Stock Analysis", "AZ"): dict(format="pct", thresholds=[
        (None, 0.20, "green", "Stronger, tangible balance sheet"),
        (0.20, 0.50, "amber", "Balanced mix"),
        (0.50, None, "red", "Heavy reliance on intangibles"),
    ]),
    ("Stock Analysis", "BA"): dict(format="x", thresholds=[
        (None, 1.0, "green", "Undervalued"),
        (1.0, 2.0, "amber", "Fair value"),
        (2.0, None, "red", "Expensive"),
    ]),
    ("Stock Analysis", "BB"): dict(format="pct", thresholds=[
        (None, 0.05, "red", "Weak profitability"),
        (0.05, 0.10, "amber", "Moderate efficiency"),
        (0.10, None, "green", "Strong, high quality"),
    ]),
    ("Stock Analysis", "BC"): dict(format="pct", thresholds=[
        (None, 0.10, "red", "Weak / poor capital use"),
        (0.10, 0.15, "amber", "Decent"),
        (0.15, None, "green", "Strong shareholder returns"),
    ]),
    ("Stock Analysis", "BD"): dict(format="pct", thresholds=[
        (None, 0.10, "red", "Weak cost control"),
        (0.10, 0.15, "amber", "Healthy, not exceptional"),
        (0.15, None, "green", "Strong pricing power"),
    ]),
    ("Stock Analysis", "BE"): dict(format="x", thresholds=[
        (None, 15, "green", "Attractive"),
        (15, 25, "amber", "Fair valuation"),
        (25, None, "red", "Expensive"),
    ]),
    ("Stock Analysis", "BF"): dict(format="pct", thresholds=[
        (None, 0.10, "red", "Weak capital efficiency"),
        (0.10, 0.15, "amber", "Acceptable"),
        (0.15, None, "green", "Excellent capital allocation"),
    ]),
    ("Stock Analysis", "BG"): dict(format="pct", thresholds=[
        (None, 0.40, "green", "Low financial risk"),
        (0.40, 0.60, "amber", "Moderate leverage"),
        (0.60, None, "red", "High reliance on debt"),
    ]),
    ("Stock Analysis", "BH"): dict(format="x", thresholds=[
        (None, 0.7, "red", "Potential short-term cash risk"),
        (0.7, 1.0, "amber", "Acceptable but tight"),
        (1.0, None, "green", "Strong immediate liquidity"),
    ]),
    ("Stock Analysis", "BI"): dict(format="x", thresholds=[
        (None, 1.0, "red", "Liquidity risk"),
        (1.0, 1.5, "amber", "Adequate"),
        (1.5, None, "green", "Good liquidity"),
    ]),
    ("Stock Analysis", "BJ"): dict(format="x", thresholds=[
        (None, 1.0, "green", "Conservative financing"),
        (1.0, 1.5, "amber", "Balanced"),
        (1.5, None, "red", "Aggressive leverage"),
    ]),
    ("Stock Analysis", "CG"): dict(format="num"),
    ("Stock Analysis", "CH"): dict(format="x"),
    # CJ "Variance and Standard Deviation" intentionally dropped (Andrew's
    # call - the column mixes two stats into one text-ish cell and isn't
    # useful as a card); see EXCLUDE_COLUMNS below.

    ("IV2BV", "AP"): dict(format="x", thresholds=[
        (None, 1.5, "red", "Weak value creation"),
        (1.5, 3.0, "amber", "Moderate value creation"),
        (3.0, None, "green", "Strong economic value"),
    ]),

    ("Dividend Ratio", "AU"): dict(format="x"),
    ("Dividend Ratio", "CB"): dict(format="cur"),

    ("Earnings Analysis", "Z"): dict(format="pct"),
    ("Earnings Analysis", "AC"): dict(format="x"),
    ("Earnings Analysis", "AD"): dict(format="x"),
    ("Earnings Analysis", "AV"): dict(format="x"),
    ("Earnings Analysis", "AW"): dict(format="pct"),
    ("Earnings Analysis", "AX"): dict(format="num", thresholds=[
        (None, 20, "green", "Cheap on both earnings & assets"),
        (20, 60, "amber", "Normal / fair valuation"),
        (60, None, "red", "Expensive (double premium)"),
    ]),

    ("Cost of Capital Analysis", "AF"): dict(format="pct"),

    ("Valuation", "Q"): dict(format="x"),
    ("Valuation", "Z"): dict(format="cur"),
    ("Valuation", "AA"): dict(format="cur"),
    ("Valuation", "AB"): dict(format="pct"),
    ("Valuation", "AO"): dict(format="x"),
    ("Valuation", "BK"): dict(format="pct"),
    ("Valuation", "BP"): dict(format="pct"),
    ("Valuation", "DE"): dict(format="cur"),
}

# Company Potential is Munger/Buffett-style qualitative judgment (free text
# + a handful of Low/Medium/High ratings), not chartable market data - kept
# in the dropdown per Andrew's instruction, content is a later batch.

# Columns that DO have an Excel comment (so they'd normally be picked up
# automatically) but that Andrew asked to drop from the app entirely.
EXCLUDE_COLUMNS = {
    ("Stock Analysis", "CJ"),  # "Variance and Standard Deviation" - not useful as a card
}

# Andrew's follow-up: for the three "thin" sheets that only had 1-2
# commented columns (Value vs Book, Retained Earnings, Cost of Capital), also show
# the sheet's other headline numbers as plain (uncoloured) cards - he never
# wrote a comment/threshold on these, but they're clearly main data (share
# price, market cap, WACC's building blocks, etc.), not incidental. Where a
# header repeats across several years (Andrew keeps TTM + a few historical
# years side by side), only the first/most-recent (TTM) column is used here
# - the historical repeats feed the dedicated charts below instead.
#
# Stock Analysis (Fundamentals) follow-up: Andrew asked for Income Tax
# Expense and % Income Paid on Taxes to be added here too - neither had an
# Excel comment, but they're clearly main data alongside the rest of the
# "Other metrics" grid.
PLAIN_EXTRA = {
    "Stock Analysis": [
        ("AE", "cur"), ("AF", "pct"),
    ],
    "IV2BV": [
        ("F", "cur"), ("G", "cur"), ("H", "cur"), ("I", "cur"),
    ],
    "Dividend Ratio": [
        ("F", "cur"), ("G", "cur"), ("H", "cur"), ("I", "cur"),
        # BQ ("Retained Earnings (TTM)" = EPS - Dividend) is genuinely
        # currency-formatted ($0.00) in the workbook, not a percent - it
        # was mistagged "pct" here (importer error, not a workbook fact),
        # which the auto engine then inherited unquestioned. Fixed to
        # "cur" so future workbook rebuilds import it correctly.
        ("Y", "cur"), ("AJ", "cur"), ("BF", "pct"), ("BQ", "cur"),
    ],
    "Cost of Capital Analysis": [
        ("I", "cur"), ("M", "cur"), ("Q", "cur"), ("U", "cur"),
        ("Y", "pct"), ("AC", "cur"), ("AJ", "cur"),
    ],
}

# Fundamentals follow-up: Andrew asked for a share price growth chart. The
# workbook already computes this itself (Stock Analysis BV:CE, one column
# per year 2026 down to 2017) as (this year's avg price - last year's avg
# price) / last year's avg price - reusing it rather than deriving our own
# growth figure from the monthly Price Calc series.
SHARE_PRICE_GROWTH_COLS = [get_column_letter(c) for c in range(74, 84)]  # BV..CE

# Earnings Trends follow-up: two more reference lines for the PE Ratio by
# Year chart - "PE ratio (Current price to 3y average EPS)" (AV, already a
# plain card too) and the plain "PE Ratio Average" (AU, which had no Excel
# comment so was never shown anywhere before).
PE_RATIO_REF_COLS = {"avg_3y": "AV", "overall_avg": "AU"}

# Cost of Capital follow-up: WACC vs ROIC per period. Andrew's workbook
# tracks both across the same 4 periods (TTM, and 3 prior years) side by
# side - reusing those columns rather than inventing a different year set.
WACC_ROIC_COLS = {
    "wacc": ["AF", "AG", "AH", "AI"],
    "roic": ["Y", "Z", "AA", "AB"],
}

# Fair Value follow-up: the key inputs behind each valuation method, shown
# under its bar. Traced from the same formulas as VALUATION_METHOD_COLS
# above (see that comment for the full derivation):
#   - PE Forward (J) = AA (Forecast EPS 5y) x AP (Actual P/E)
#   - PE Trailing (K) = AQ (EPS Diluted, i.e. trailing EPS) x AO (Average P/E)
#   - DCF (P) uses BO (Perpetual Rate) and BP (Owner Earnings Growth Rate,
#     i.e. the base-case growth assumption fed into the 10y FCF projection)
#   - Rational Compounder Method 10y (U, renamed from "Equity Method 10y"
#     per Andrew's request) uses CR (Equity Growth) and CT (the discount
#     rate used specifically in this calc, separate from BK's DCF discount
#     rate)
VALUATION_INPUT_COLS = {
    "pe_forward": [("Forecast EPS (5y)", "AA", "cur"), ("Actual P/E", "AP", "x")],
    "pe_trailing": [("EPS (Trailing/Diluted)", "AQ", "cur"), ("Average P/E", "AO", "x")],
    "dcf": [("Discount Rate", "BK", "pct"), ("Perpetual Rate", "BO", "pct"),
            ("Base Case Growth", "BP", "pct")],
    "equity_10y": [("Equity Growth", "CR", "pct"), ("Discount Rate (this calc)", "CT", "pct")],
}

# Company Potential follow-up: Andrew's free-text answers, grammar/spelling
# corrected while keeping his own writing voice (no rewrites, no content
# changes - see company_potential_corrections_log.txt for the original
# one-off pass). Keyed the same way as the workbook: {ticker: {column:
# {label, source_hash, text}}}. source_hash is a hash of the RAW text this
# correction was made from -- lets a later build tell "still the same
# answer, reuse the cached correction for free" apart from "Andrew edited
# this cell, needs a fresh AI pass" without re-billing for unchanged text
# every single rebuild. Entries from before this hash existed (the
# original AUB.AX/CSL.AX pass) are grandfathered in as trusted rather than
# re-corrected once, then get a hash stamped on first use so staleness
# detection works from then on.
CP_CORRECTIONS_PATH = os.path.join(_cp_data_dir(), "company_potential_corrections.json")

_CP_CORRECTION_MODEL = "claude-haiku-4-5"
_CP_CORRECTION_SYSTEM_PROMPT = (
    "You are a careful copy editor reviewing an individual investor's own "
    "private research notes on a stock. For each item in the JSON array "
    "below, fix ONLY clear spelling and grammar mistakes. Do NOT rewrite, "
    "rephrase, restructure, summarize, or change the meaning, opinions, "
    "conclusions, tone, or personal writing voice -- preserve the "
    "author's own style exactly, including any informal phrasing, as "
    "long as it's grammatically valid. If an item already has no errors, "
    "return it completely unchanged, word for word. Never add "
    "commentary, disclaimers, or anything not in the original text.\n\n"
    "Return ONLY a JSON array, the same length and order as the input, "
    "where each element is {\"key\": <key>, \"text\": <corrected text>}. "
    "No markdown code fences, no text before or after the JSON."
)


def _cp_text_hash(text):
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def _cp_load_corrections_cache():
    _cp_seed_from_repo_if_missing("company_potential_corrections.json")
    if not os.path.exists(CP_CORRECTIONS_PATH):
        return {}
    with open(CP_CORRECTIONS_PATH) as f:
        return json.load(f)


def _cp_save_corrections_cache(cache):
    with open(CP_CORRECTIONS_PATH, "w") as f:
        json.dump(cache, f, indent=1)


def _cp_correct_batch_via_ai(items, api_key, ticker, warnings):
    """items: [{"key", "label", "text"}, ...] for ONE ticker's free-text
    answers that need a fresh check (new or edited since the last cached
    correction). One API call per ticker, not per paragraph, to keep this
    cheap. Returns {key: corrected_text} -- falls back to each item's own
    raw text (never crashes the whole rebuild) if the package is missing,
    the call fails, or the model's response doesn't parse cleanly."""
    fallback = {it["key"]: it["text"] for it in items}
    try:
        import anthropic
    except ImportError:
        warnings.append(
            "Grammar check skipped: the 'anthropic' package isn't installed "
            "(add it to requirements.txt) - showing raw text instead."
        )
        return fallback
    try:
        client = anthropic.Anthropic(api_key=api_key)
        payload = [{"key": it["key"], "text": it["text"]} for it in items]
        resp = client.messages.create(
            model=_CP_CORRECTION_MODEL,
            max_tokens=8192,
            system=_CP_CORRECTION_SYSTEM_PROMPT,
            messages=[{"role": "user", "content": json.dumps(payload)}],
        )
        raw = "".join(
            block.text for block in resp.content if getattr(block, "type", None) == "text"
        ).strip()
        if raw.startswith("```"):
            raw = raw.strip("`")
            if raw.lower().startswith("json"):
                raw = raw[4:]
            raw = raw.strip()
        parsed = json.loads(raw)
        result = dict(fallback)
        for entry in parsed:
            key = entry.get("key") if isinstance(entry, dict) else None
            text = entry.get("text") if isinstance(entry, dict) else None
            if key in result and isinstance(text, str) and text.strip():
                result[key] = text
        return result
    except Exception as exc:
        warnings.append(
            f"Grammar check failed for {ticker} ({exc}) - showing raw text instead."
        )
        return fallback


def _cp_apply_corrections(tkr, raw_items, cache, api_key, warnings):
    """raw_items: this ticker's free-text answers for the CURRENT build.
    Returns {key: text_to_use}, reusing a cached correction whenever the
    raw text's hash still matches what that correction was made from, and
    only calling the AI (one batched call) for genuinely new/changed text.
    With no api_key configured, new/changed text is left exactly as typed
    -- this feature is fully opt-in, same as the rest of the app's
    optional integrations."""
    ticker_cache = cache.get(tkr, {})
    result = {}
    to_correct = []
    for it in raw_items:
        h = _cp_text_hash(it["text"])
        cached_entry = ticker_cache.get(it["key"])
        if cached_entry and cached_entry.get("source_hash") == h:
            result[it["key"]] = cached_entry["text"]
        elif cached_entry and "source_hash" not in cached_entry:
            result[it["key"]] = cached_entry["text"]
            ticker_cache[it["key"]] = {
                "label": it["label"], "source_hash": h, "text": cached_entry["text"],
            }
        else:
            to_correct.append(it)

    if to_correct:
        if not api_key:
            for it in to_correct:
                result[it["key"]] = it["text"]
        else:
            corrected = _cp_correct_batch_via_ai(to_correct, api_key, tkr, warnings)
            for it in to_correct:
                text = corrected.get(it["key"], it["text"])
                result[it["key"]] = text
                ticker_cache[it["key"]] = {
                    "label": it["label"], "source_hash": _cp_text_hash(it["text"]), "text": text,
                }

    if ticker_cache:
        cache[tkr] = ticker_cache
    return result


# Value vs Book: Andrew asked for Intrinsic Value vs Book Value shown as an
# actual bar comparison (replacing the single IV/BV ratio gauge) so the
# value-creation gap is visible, not just a compressed multiple. TTM columns.
IV_BV_BAR_COLS = {"price": "F", "iv": "T", "bv": "AE"}

# Andrew's follow-up: IV/BV as a proper X=year, Y=IV/BV chart across every
# year he's modelled (AP=TTM down to AZ=2016), not just the current ratio.
IV_BV_SERIES_COLS = [
    "AP", "AQ", "AR", "AS", "AT", "AU", "AV", "AW", "AX", "AY", "AZ",
]

# Earnings Trends follow-up: three more year-series charts (EPS itself,
# EPS growth, and PE Ratio), plus a batch of plain summary figures that
# never had a comment but are clearly "main data" (10y/4y average earnings,
# max/min earnings, the SD figures, the two growth-rate summaries, and the
# plain PE Ratio Average). AC/AD (10y/4y AVG+SD) and AV/AW/AX stay handled
# by METRIC_META above as before.
EARNINGS_SERIES = {
    "eps": [get_column_letter(c) for c in range(11, 22)],       # K..U
    "eps_growth": [get_column_letter(c) for c in range(31, 41)],  # AE..AN
    "pe_ratio": [get_column_letter(c) for c in range(43, 47)],    # AQ..AT
}
EARNINGS_PLAIN_EXTRA = [
    ("V", "cur"), ("W", "cur"), ("X", "cur"), ("Y", "cur"),
    ("Z", "num"), ("AA", "cur"), ("AB", "cur"),
    ("AO", "pct"), ("AP", "pct"), ("AU", "x"),
]
PLAIN_EXTRA["Earnings Analysis"] = EARNINGS_PLAIN_EXTRA

# Fair Value follow-up: Andrew wants ONLY these four intrinsic-value methods
# (replacing the previous general metrics grid), compared against current
# price. None of the four have a comment in Excel, so column identity was
# worked out from the formulas themselves:
#   - PE Method (Forward) / (Trailing): J / K, direct EPS x multiple.
#   - "DCF": P "Buffet Share Price" - traced its formula back through CB
#     "Share Price Terminal" = SUM of BQ:CA (10 years of discounted Free
#     Cash Flow, columns "Discount Value Year 1..10", plus a terminal
#     value) / shares outstanding - a genuine discounted-cash-flow model.
#     (DE "Share Price 10Y" looked like a DCF candidate at first, but its
#     formula turned out to be the SAME calculation as U below, just before
#     currency conversion - not a distinct 4th method, so it's not used
#     here.)
#   - "My equity method 10y": U "Equity Method 10y" - a separate model
#     based on projected equity + earnings growth (see its own comment on
#     DE, which U is currency-converted from).
VALUATION_METHOD_COLS = {
    "price": "H", "pe_forward": "J", "pe_trailing": "K",
    "dcf": "P", "equity_10y": "U",
}

# Retained Earnings: Andrew's own workbook already computes a genuine "Value
# Created" test (Ben Graham/Buffett style: for every $1 of earnings
# retained, how much market cap value did that create) at four horizons.
# Never had a comment on it, but it's exactly the "value created" chart he
# asked for - surface it as a grouped bar (retained-earnings put in vs
# value created) rather than a plain metric card.
VALUE_CREATED_HORIZONS = [
    ("2Y", "CH", "CJ"),
    ("5Y", "CE", "CG"),
    ("10Y", "CB", "CD"),
    ("TTM", "CK", "CM"),
]

# ---- Company Potential follow-up ----
# Andrew: "the cells that have a drop down selection of high medium or low
# to be called directly" - these are the columns whose intended values are
# literally Low/Medium/High. "polarity" is only used to pick a colour and
# is a judgment call (documented in the README, not something Andrew
# specified per-column): good_high = High is green/good, good_low = Low is
# green/good, neutral = shown as a plain badge, no red/green judgment.
#
# Every entry is (fallback_letter, raw_header_text, display_label, ...).
# raw_header_text is the actual current header text from Andrew's sheet
# (typos and all) -- at build time this is used to FIND the column by its
# header, wherever it currently sits, so inserting/reordering columns
# doesn't silently misattribute data the way a bare fixed-letter lookup
# would. fallback_letter (today's known position) is only used if that
# header text can't be found, or matches more than one column -- see
# _cp_resolve_column(). display_label is what the app actually shows.
COMPANY_POTENTIAL_HML = [
    ("O", "Management Reputation", "Management Reputation", "good_high"),
    ("P", "Debt  Exposure", "Debt Exposure", "good_low"),
    ("Q", "Legal Exposure", "Legal Exposure", "good_low"),
    ("R", "Inflation Exposure", "Inflation Exposure", "good_low"),
    ("S", "Business Understanding", "Business Understanding", "good_high"),
    ("T", "Value Prospect", "Value Prospect", "good_high"),
    ("U", "Progress Prospect", "Progress Prospect", "good_high"),
    ("V", "Weath Prospect", "Wealth Prospect", "good_high"),
    ("AE", "Market Sentiment", "Market Sentiment", "neutral"),
    ("AK", "Risk", "Risk", "good_low"),
    ("AL", "Insights", "Insights", "good_high"),
    ("BC", "Stability of the Compnay within Industry. Quantity Vs Qualitative",
     "Stability within Industry", "good_high"),
    ("BL", "Do they have the ability to change product price without losing clients?",
     "Ability to Change Pricing", "good_high"),
    ("BP", "Market Activity", "Market Activity", "good_high"),
]

# Short Yes/No (or similarly terse) checks - kept separate from the H/M/L
# ratings above (different vocabulary) and from the long free-text answers
# below (these are one to few words, not paragraphs).
COMPANY_POTENTIAL_YESNO = [
    ("Y", "Is there public interest?", "Public interest?"),
    ("AC", "Ok when Inverted?", "OK when inverted?"),
    ("BB", "Great Company on a fair price", "Great company at a fair price?"),
    ("BF", "Any Shares Buyback/Shares Dilution?", "Any share buybacks?"),
    ("BH", "True Earnings?", "True earnings?"),
    ("BK", "Are fixed charges (i.e cost of production, bonds, preferred shares etc) "
     "high so it will magnify the earnings when increasing?", "High fixed charges?"),
    ("BM", "Will I Hold the Stock Through Bear Market", "Would hold through a bear market?"),
    ("BN", "Are the forecasted earnings possible, plausible and probable?",
     "Forecasted earnings possible/plausible/probable?"),
]

# "the wants with text to be merge in a few text boxes with the title and
# content that you see fit for the grouping" - Andrew's own column headers
# grouped into a handful of themed sections (my grouping call, not his).
# NOTE: AZ and BA are both literally headed "Tendencies Company 2" in the
# source sheet (Andrew's own duplicate, not a typo on my end) - header text
# alone can't tell them apart, so both fall back to their fixed letters and
# _cp_resolve_column() logs a warning about it rather than guessing.
COMPANY_POTENTIAL_TEXT_GROUPS = [
    ("The Business & Its Moat", [
        ("AG", "Market Reality", "Market Reality"),
        ("AH", "Challenges", "Challenges"),
        ("AP", "Tradeamrk Product", "Trademark Product / Differentiator"),
        ("AR", "Advanatge of Scale with no Burocracy and momentum?", "Advantage of Scale"),
        ("AS", "Is it specialised within its ecosystem? Does it have a good "
         "reputation in this secialization? Is it a predator?",
         "Specialised Within Its Ecosystem"),
        ("AT", "Circle of Competence", "Circle of Competence"),
        ("AW", "Big Wave to Ride", "Big Wave to Ride"),
        ("BO", "Source of Income?", "Source of Income"),
        ("BR", "Small and Promising or large and ordinary?",
         "Small & Promising, or Large & Ordinary"),
    ]),
    ("Risk & Inversion", [
        ("AB", "Invertion Angle", "Inversion Angle"),
        ("AF", "Breakeven Price", "Breakeven Price"),
        ("AI", "No Brainer Question for this Company?", "No Brainer Question"),
        ("AJ", "1 Foot Fence With Big Rewards on the other side?",
         "1 Foot Fence, Big Reward the Other Side"),
        ("AU", "Probability and decision investment trees", "Probability & Decision Trees"),
    ]),
    ("Psychology & Munger Tendencies", [
        ("AM", "Sauerkaraut effect?", "Sauerkraut Effect"),
        ("AN", "Lollapalooza effect?", "Lollapalooza Effect"),
        ("AO", "Ideology affected?", "Ideology Affected"),
        ("AQ", "What is Multidisciplinary Approach for this Company?ID the "
         "Disciplines that particpate in this business and its impact.",
         "Multidisciplinary Approach"),
        ("AV", "First Principles", "First Principles"),
        ("AX", "Tendencies Market/Product 1", "Tendencies - Market/Product 1"),
        ("AY", "Tendencies Market/Product 2", "Tendencies - Market/Product 2"),
        ("AZ", "Tendencies Company 2", "Tendencies - Company 1"),
        ("BA", "Tendencies Company 2", "Tendencies - Company 2"),
    ]),
    ("The Investment Case", [
        ("BD", "Why is a Good Investment? What is advantage over the others?",
         "Why Is This a Good Investment?"),
        ("BX", "Buffet Tenets", "Buffett Tenets"),
        ("BY", "Stock Vs Bond Comparison", "Stock vs Bond Comparison"),
        ("CA", "Which famous investor have invested in this stock?",
         "Which Famous Investors Hold This?"),
        ("BZ", "Have you done a chart analysis to determine entry and exit?", "Chart Analysis"),
        ("BV", "Investment Recommendation", "Investment Recommendation"),
        ("BS", "Speculation/Investment", "Speculation / Investment Type"),
        ("BQ", "Popular?", "Popular?"),
    ]),
    ("Financial Diligence", [
        ("BE", "Have you Checked the Taxes? Do they make sense?", "Have You Checked the Taxes?"),
        ("BJ", "Earning Forecast", "Earning Forecast"),
        ("BG", "If it were a Private Business How Will I Measure it?",
         "If a Private Business, How Would I Measure It?"),
    ]),
    ("Management & Context", [
        ("W", "CEO Time with Company", "CEO Time With Company"),
        ("X", "How old is the Business?", "How Old Is the Business?"),
        ("Z", "Competitors Analysis 1", "Competitors Analysis 1"),
        ("AA", "Competitors Analysis 2", "Competitors Analysis 2"),
        ("BW", "Empoyees", "Employees"),
    ]),
]


import re as _cp_re


def _cp_normalize_header(text):
    """Loose match: ignores case, whitespace and punctuation, so re-typing
    a header with different spacing or a trailing '?' still matches. NOT
    typo-tolerant on purpose -- a genuine rewording still falls through to
    the fallback-letter + warning path below rather than guessing."""
    return _cp_re.sub(r"[^a-z0-9]+", "", text.lower())


def _cp_header_letter_map(ws_c):
    """Scan the Company Potential header rows once per build and return
    {normalized_header_text: [column_letters]} so columns below can be
    resolved by their actual current header text rather than assuming a
    fixed letter forever. A header can map to more than one letter (the
    sheet has at least one genuine duplicate) - callers decide what to do
    with that via _cp_resolve_column()."""
    header_map = {}
    for c in range(1, ws_c.max_column + 1):
        letter = get_column_letter(c)
        header_text = None
        for hr in HEADER_ROWS["Company Potential"]:
            v = ws_c.cell(row=hr, column=c).value
            if v not in (None, ""):
                header_text = str(v).strip()
        if header_text:
            header_map.setdefault(_cp_normalize_header(header_text), []).append(letter)
    return header_map


def _cp_resolve_column(header_map, fallback_letter, raw_header_text, field_label,
                        warnings, occurrence_index=0):
    """Find `raw_header_text` in the CURRENT sheet and return whichever
    column it's in now (works even if columns were inserted/reordered
    since raw_header_text was last recorded).

    When the header text is genuinely duplicated (two+ columns share the
    exact same text -- e.g. "Tendencies Company 2" appearing twice),
    `occurrence_index` picks the Nth match left-to-right instead of
    blindly trusting a fixed fallback letter, which goes stale (and can
    point at a totally unrelated column) the moment columns shift.
    `occurrence_index` is each field's position among the OTHER
    same-titled fields in config declaration order, which matches how
    those duplicate headers are laid out left-to-right in the sheet.
    Still logs a warning either way, since a rename to make the headers
    unique is the only way to remove the guesswork entirely.

    Falls back to the last-known fixed letter -- and appends a
    human-readable warning -- when the header can't be found at all, so a
    genuine rename/deletion surfaces instead of silently reading the
    wrong (or now-missing) column."""
    matches = header_map.get(_cp_normalize_header(raw_header_text), [])
    if len(matches) == 1:
        return matches[0]
    if len(matches) > 1:
        if occurrence_index < len(matches):
            resolved = matches[occurrence_index]
            warnings.append(
                f"Company Potential '{field_label}': header text matches {len(matches)} "
                f"columns ({', '.join(matches)}) - used column {resolved} (position "
                f"{occurrence_index + 1} of {len(matches)} duplicates, left to right). "
                f"Give one of the duplicate headers a unique name to remove the "
                f"guesswork."
            )
            return resolved
        warnings.append(
            f"Company Potential '{field_label}': header text matches {len(matches)} "
            f"columns ({', '.join(matches)}) - used the last-known column "
            f"{fallback_letter}. Give one of the duplicate headers a unique name "
            f"to fix this."
        )
        return fallback_letter
    warnings.append(
        f"Company Potential '{field_label}': couldn't find its header "
        f"({raw_header_text!r}) anywhere in the sheet - used the last-known column "
        f"{fallback_letter} instead. If that header was renamed or the column was "
        f"deleted, this may now be reading the wrong data - worth checking."
    )
    return fallback_letter


def _clean_num(v):
    if isinstance(v, (int, float)):
        return float(v)
    return None


def build(path, anthropic_api_key=None):
    """anthropic_api_key: enables the live Company Potential grammar/wording
    check (see _cp_correct_batch_via_ai) when set, either passed explicitly
    or via the ANTHROPIC_API_KEY environment variable. Entirely optional --
    with no key available (the default for a plain `python3
    build_compounder_data.py` run, or before it's configured on Railway),
    Company Potential text is shown exactly as typed, same as every other
    optional integration in this app (NewsAPI, Wisesheets, etc.)."""
    anthropic_api_key = (anthropic_api_key or os.environ.get("ANTHROPIC_API_KEY", "")).strip() or None
    wb_c = openpyxl.load_workbook(path, data_only=False)  # comments
    wb_v = openpyxl.load_workbook(path, data_only=True)   # cached values

    out = {"generated_at": datetime.now(timezone.utc).isoformat(), "tickers": {}, "sections": {}}

    # Ticker universe + basic identity, from Stock Analysis (widest sheet).
    ws_c = wb_c["Stock Analysis"]
    ws_v = wb_v["Stock Analysis"]
    tickers = []
    for r in range(4, min(ws_v.max_row, MAX_DATA_ROWS) + 1):
        tkr = ws_v.cell(row=r, column=4).value
        if isinstance(tkr, str) and tkr.strip() and "VALUE" not in tkr and "#" not in tkr:
            industry = ws_v.cell(row=r, column=5).value
            ceo = ws_v.cell(row=r, column=6).value
            out["tickers"][tkr] = {
                "industry": industry if isinstance(industry, str) else None,
                "ceo": ceo if isinstance(ceo, str) else None,
                "row": r,
            }
            tickers.append(tkr)

    row_for_ticker_by_sheet = {}  # cached for the price/IV-BV/value-created extras below

    for section_label, sheet_name in SECTIONS:
        wsc = wb_c[sheet_name]
        wsv = wb_v[sheet_name]
        tcol = TICKER_COL[sheet_name]
        # map ticker -> row within THIS sheet (rows line up 1:1 with Stock
        # Analysis for every sheet in this workbook, but look it up fresh
        # in case a sheet is missing a row for some ticker).
        row_for_ticker = {}
        for r in range(4, min(wsv.max_row, MAX_DATA_ROWS) + 1):
            tkr = wsv.cell(row=r, column=tcol).value
            if isinstance(tkr, str) and tkr in out["tickers"]:
                row_for_ticker[tkr] = r
        row_for_ticker_by_sheet[sheet_name] = row_for_ticker

        plain_extra_cols = {col: fmt for col, fmt in PLAIN_EXTRA.get(sheet_name, [])}

        metrics = []
        max_col = wsc.max_column
        for c in range(1, max_col + 1):
            letter = get_column_letter(c)
            if (sheet_name, letter) in EXCLUDE_COLUMNS:
                continue
            header_text, comment_text = None, None
            for hr in HEADER_ROWS[sheet_name]:
                cell = wsc.cell(row=hr, column=c)
                if cell.value not in (None, ""):
                    header_text = str(cell.value).strip()
                if cell.comment is not None:
                    comment_text = cell.comment.text.strip()
            is_plain_extra = letter in plain_extra_cols
            if comment_text is None and not is_plain_extra:
                continue  # only the columns Andrew himself annotated (+ curated extras)
            if header_text is None:
                continue
            if c == tcol:
                continue  # skip the ticker column itself

            meta = METRIC_META.get((sheet_name, letter))
            if sheet_name == "Company Potential":
                fmt = "text"
                thresholds = None
            elif meta is not None:
                fmt = meta.get("format", "num")
                thresholds = meta.get("thresholds")
            elif is_plain_extra:
                fmt, thresholds = plain_extra_cols[letter], None
                # disambiguate repeated headers (e.g. 4 "Market Cap" columns)
                # with the period row Andrew put above the header (TTM, 2025...)
                period = wsv.cell(row=2, column=c).value
                if period not in (None, ""):
                    header_text = f"{header_text} ({period})"
            else:
                fmt, thresholds = "num", None

            values = {}
            for tkr, r in row_for_ticker.items():
                raw = wsv.cell(row=r, column=c).value
                if fmt == "text":
                    values[tkr] = raw if isinstance(raw, str) and raw.strip() else None
                else:
                    values[tkr] = _clean_num(raw)

            metrics.append({
                "key": letter,
                "label": header_text,
                "comment": comment_text or "",
                "format": fmt,
                "thresholds": thresholds,
                "values": values,
            })

        out["sections"][section_label] = {"sheet": sheet_name, "metrics": metrics}

    # ---- Fundamentals: price history + Andrew's own "10y Average" line ----
    sa_rows = row_for_ticker_by_sheet["Stock Analysis"]
    ws_price = wb_v["Price Calc"]
    price_history = {}
    for tkr, sa_row in sa_rows.items():
        pc_row = sa_row - 1  # Price Calc rows are Stock Analysis rows, offset by 1
        dates, prices = [], []
        for c in range(2, 123):  # B..DR - the monthly price columns
            d = ws_price.cell(row=2, column=c).value
            p = ws_price.cell(row=pc_row, column=c).value
            p = _clean_num(p)
            if d is not None and p is not None:
                dates.append(d.isoformat() if hasattr(d, "isoformat") else str(d))
                prices.append(p)
        avg_10y = _clean_num(ws_v.cell(row=sa_row, column=column_index_from_string("CF")).value)
        if dates:
            price_history[tkr] = {"dates": dates, "prices": prices, "avg_10y": avg_10y}
    out["sections"]["Fundamentals"]["price_history"] = price_history

    # ---- Fundamentals follow-up: Share Price Growth by year ----
    share_price_growth = {}
    for tkr, sa_row in sa_rows.items():
        years, values_ = [], []
        for col in SHARE_PRICE_GROWTH_COLS:
            c = column_index_from_string(col)
            period = ws_v.cell(row=3, column=c).value
            val = _clean_num(ws_v.cell(row=sa_row, column=c).value)
            if period not in (None, "") and val is not None:
                years.append(str(period))
                values_.append(val)
        if years:
            share_price_growth[tkr] = {"years": years, "values": values_}
    out["sections"]["Fundamentals"]["share_price_growth"] = share_price_growth

    # ---- Value vs Book: IV vs BV vs current price, as an actual comparison
    # (Andrew asked for this to replace the single IV/BV ratio gauge) ----
    iv2bv_rows = row_for_ticker_by_sheet.get("IV2BV", {})
    ws_iv2bv = wb_v["IV2BV"]
    iv_bv_bar = {}
    for tkr, r in iv2bv_rows.items():
        entry = {}
        for key, col in IV_BV_BAR_COLS.items():
            entry[key] = _clean_num(ws_iv2bv.cell(row=r, column=column_index_from_string(col)).value)
        if any(v is not None for v in entry.values()):
            iv_bv_bar[tkr] = entry
    out["sections"]["Value vs Book"]["iv_bv_bar"] = iv_bv_bar

    # ---- Value vs Book follow-up: IV/BV across every modelled year
    # (X=year, Y=IV/BV), not just the current ratio ----
    iv_bv_series = {}
    for tkr, r in iv2bv_rows.items():
        years, ratios = [], []
        for col in IV_BV_SERIES_COLS:
            c = column_index_from_string(col)
            period = ws_iv2bv.cell(row=2, column=c).value
            ratio = _clean_num(ws_iv2bv.cell(row=r, column=c).value)
            if period not in (None, "") and ratio is not None:
                years.append(str(period))
                ratios.append(ratio)
        if years:
            iv_bv_series[tkr] = {"years": years, "ratios": ratios}
    out["sections"]["Value vs Book"]["iv_bv_series"] = iv_bv_series

    # ---- Earnings Trends follow-up: EPS, EPS Growth, and PE Ratio across
    # every modelled year ----
    ea_rows = row_for_ticker_by_sheet.get("Earnings Analysis", {})
    ws_ea = wb_v["Earnings Analysis"]
    earnings_series = {}
    for tkr, r in ea_rows.items():
        series_out = {}
        for series_key, cols in EARNINGS_SERIES.items():
            years, values_ = [], []
            for col in cols:
                c = column_index_from_string(col)
                period = ws_ea.cell(row=2, column=c).value
                val = _clean_num(ws_ea.cell(row=r, column=c).value)
                if period not in (None, "") and val is not None:
                    years.append(str(period))
                    values_.append(val)
            if years:
                series_out[series_key] = {"years": years, "values": values_}
        if series_out:
            earnings_series[tkr] = series_out
    out["sections"]["Earnings Trends"]["series"] = earnings_series

    # ---- Earnings Trends follow-up: PE ratio reference lines (current
    # price / 3y average EPS, and the plain PE Ratio Average) for the PE
    # Ratio by Year chart ----
    pe_ratio_refs = {}
    for tkr, r in ea_rows.items():
        entry = {}
        for key, col in PE_RATIO_REF_COLS.items():
            entry[key] = _clean_num(ws_ea.cell(row=r, column=column_index_from_string(col)).value)
        if any(v is not None for v in entry.values()):
            pe_ratio_refs[tkr] = entry
    out["sections"]["Earnings Trends"]["pe_ratio_refs"] = pe_ratio_refs

    # ---- Fair Value follow-up: ONLY the 4 intrinsic-value methods Andrew
    # asked for, compared against current price ----
    val_rows = row_for_ticker_by_sheet.get("Valuation", {})
    ws_val = wb_v["Valuation"]
    valuation_methods = {}
    for tkr, r in val_rows.items():
        entry = {}
        for key, col in VALUATION_METHOD_COLS.items():
            entry[key] = _clean_num(ws_val.cell(row=r, column=column_index_from_string(col)).value)
        if any(v is not None for v in entry.values()):
            valuation_methods[tkr] = entry
    out["sections"]["Fair Value"]["valuation_methods"] = valuation_methods

    # ---- Fair Value follow-up: the key inputs behind each method, to show
    # under its bar (see VALUATION_INPUT_COLS above for the derivation) ----
    valuation_inputs = {}
    for tkr, r in val_rows.items():
        entry = {}
        for method_key, inputs in VALUATION_INPUT_COLS.items():
            items = []
            for label, col, fmt in inputs:
                val = _clean_num(ws_val.cell(row=r, column=column_index_from_string(col)).value)
                if val is not None:
                    items.append({"label": label, "value": val, "format": fmt})
            if items:
                entry[method_key] = items
        if entry:
            valuation_inputs[tkr] = entry
    out["sections"]["Fair Value"]["valuation_inputs"] = valuation_inputs

    # ---- Cost of Capital follow-up: WACC vs ROIC per period, as a grouped
    # bar chart ----
    coc_rows = row_for_ticker_by_sheet.get("Cost of Capital Analysis", {})
    ws_coc = wb_v["Cost of Capital Analysis"]
    wacc_roic_series = {}
    for tkr, r in coc_rows.items():
        series_out = {}
        for series_key, cols in WACC_ROIC_COLS.items():
            periods, values_ = [], []
            for col in cols:
                c = column_index_from_string(col)
                period = ws_coc.cell(row=2, column=c).value
                val = _clean_num(ws_coc.cell(row=r, column=c).value)
                if period not in (None, "") and val is not None:
                    periods.append(str(period))
                    values_.append(val)
            if periods:
                series_out[series_key] = {"periods": periods, "values": values_}
        if series_out:
            wacc_roic_series[tkr] = series_out
    out["sections"]["Cost of Capital"]["wacc_roic_series"] = wacc_roic_series

    # ---- Company Potential follow-up: H/M/L ratings called out directly,
    # short Yes/No checks, and the long free-text answers merged into a
    # handful of themed groups ----
    cp_rows = row_for_ticker_by_sheet.get("Company Potential", {})
    ws_cp = wb_v["Company Potential"]
    ws_cp_c = wb_c["Company Potential"]

    # Excel error strings as openpyxl (data_only) returns them - the ONLY
    # things this guard should skip. The old check ('"VALUE" not in v')
    # silently discarded any answer CONTAINING the word VALUE anywhere -
    # e.g. AUB's Challenges cell, which has an all-caps note ending in
    # "...LIMIT VALUE IS 8.5" and therefore never imported.
    _XL_ERRORS = ("#VALUE!", "#REF!", "#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!")

    def _cp_text(r, col):
        v = ws_cp.cell(row=r, column=column_index_from_string(col)).value
        if isinstance(v, str):
            t = v.strip()
            if t and t not in _XL_ERRORS:
                return t
        return None

    # Resolve every configured field to its CURRENT column once per build
    # (header text -> letter, falling back to the last-known letter with a
    # warning - see _cp_resolve_column) rather than trusting the fixed
    # letters blindly. This is what makes the sheet tolerant of Andrew
    # inserting/reordering columns in Company Potential: as long as a
    # field's header text is still somewhere in the sheet, its answers keep
    # showing up under the right label even if it moved.
    build_warnings = out.setdefault("build_warnings", [])
    header_map = _cp_header_letter_map(ws_cp_c)

    # Shared across all three tables (in declaration order) so that if the
    # same header text is duplicated, each field gets a distinct position
    # among the duplicates -- see _cp_resolve_column's occurrence_index.
    _cp_occurrence_counts = {}

    def _cp_next_occurrence(raw_header_text):
        key = _cp_normalize_header(raw_header_text)
        idx = _cp_occurrence_counts.get(key, 0)
        _cp_occurrence_counts[key] = idx + 1
        return idx

    hml_resolved = [
        (_cp_resolve_column(header_map, letter, header_text, label, build_warnings,
                             _cp_next_occurrence(header_text)), label, polarity)
        for letter, header_text, label, polarity in COMPANY_POTENTIAL_HML
    ]
    yesno_resolved = [
        (_cp_resolve_column(header_map, letter, header_text, label, build_warnings,
                             _cp_next_occurrence(header_text)), label)
        for letter, header_text, label in COMPANY_POTENTIAL_YESNO
    ]
    text_groups_resolved = [
        (group_title, [
            (_cp_resolve_column(header_map, letter, header_text, label, build_warnings,
                                 _cp_next_occurrence(header_text)), label)
            for letter, header_text, label in cols
        ])
        for group_title, cols in COMPANY_POTENTIAL_TEXT_GROUPS
    ]

    hml_ratings = {}
    for tkr, r in cp_rows.items():
        ratings = []
        for col, label, polarity in hml_resolved:
            val = _cp_text(r, col)
            if val:
                ratings.append({"key": col, "label": label, "value": val, "polarity": polarity})
        if ratings:
            hml_ratings[tkr] = ratings
    out["sections"]["Company Potential"]["hml_ratings"] = hml_ratings

    yesno_checks = {}
    for tkr, r in cp_rows.items():
        checks = []
        for col, label in yesno_resolved:
            val = _cp_text(r, col)
            if val:
                checks.append({"key": col, "label": label, "value": val})
        if checks:
            yesno_checks[tkr] = checks
    out["sections"]["Company Potential"]["yesno_checks"] = yesno_checks

    # Grammar/wording correction (spelling + grammar only, voice/content
    # untouched - see _cp_correct_batch_via_ai's system prompt) runs here,
    # ONE batched AI call per ticker covering every free-text answer that's
    # new or has changed since the last build; anything unchanged reuses
    # its cached correction for free. See company_potential_corrections_log.txt
    # for the original one-off pass this cache started from.
    cp_corrections_cache = _cp_load_corrections_cache()
    text_groups = {}
    for tkr, r in cp_rows.items():
        raw_by_col = {}
        for group_title, cols in text_groups_resolved:
            for col, label in cols:
                val = _cp_text(r, col)
                if val:
                    raw_by_col[col] = {"key": col, "label": label, "text": val}
        if not raw_by_col:
            continue
        corrected_by_col = _cp_apply_corrections(
            tkr, list(raw_by_col.values()), cp_corrections_cache,
            anthropic_api_key, build_warnings,
        )
        groups = []
        for group_title, cols in text_groups_resolved:
            items = []
            for col, label in cols:
                if col in raw_by_col:
                    text = corrected_by_col.get(col, raw_by_col[col]["text"])
                    items.append({"key": col, "label": label, "text": text})
            if items:
                groups.append({"title": group_title, "items": items})
        if groups:
            text_groups[tkr] = groups
    _cp_save_corrections_cache(cp_corrections_cache)
    out["sections"]["Company Potential"]["text_groups"] = text_groups

    # ---- Retained Earnings: Andrew's own "Value Created" retained-earnings
    # test, at 2Y/5Y/10Y/TTM horizons - never had a comment, but it's
    # exactly the "value created" chart he asked for. ----
    div_rows = row_for_ticker_by_sheet.get("Dividend Ratio", {})
    ws_div = wb_v["Dividend Ratio"]
    value_created = {}
    for tkr, r in div_rows.items():
        horizons = {}
        for label, retained_col, created_col in VALUE_CREATED_HORIZONS:
            retained = _clean_num(ws_div.cell(row=r, column=column_index_from_string(retained_col)).value)
            created = _clean_num(ws_div.cell(row=r, column=column_index_from_string(created_col)).value)
            if retained is not None or created is not None:
                horizons[label] = {"retained_earnings": retained, "value_created": created}
        if horizons:
            value_created[tkr] = horizons
    out["sections"]["Retained Earnings"]["value_created"] = value_created

    return out


# ---- Merge-not-overwrite on re-upload: REMOVED (audit fix 2.11) ----
#
# This used to be where merge_compounder_data() + its helpers
# (_cp_is_empty/_cp_list_item_id/_cp_merge_value) lived - built to stop a
# fresh/leaner spreadsheet re-upload from silently deleting anything not
# re-typed into the new file. The admin rebuild flow (app.py) deliberately
# moved to a full replace instead: "the workbook you upload is exactly
# what the site shows, and history lives in the archive instead of
# leaking into the present" (see app.py's rebuild handler, and
# archive_current_snapshot()/list_archived_snapshots() below, which are
# the actual safety net now - every rebuild is archived first and
# browsable via the Research page's snapshot picker). merge_compounder_data()
# was never called from that path (or anywhere else) once that decision
# was made, leaving ~85 lines of dead code whose comments described
# protection that wasn't actually active - removed rather than left to
# mislead the next reader. Recoverable from git history if a future
# rebuild flow ever wants merge-on-upload back.


ARCHIVE_DIR_NAME = "compounder_archive"


def _cp_archive_dir():
    path = os.path.join(_cp_data_dir(), ARCHIVE_DIR_NAME)
    os.makedirs(path, exist_ok=True)
    return path


def archive_current_snapshot():
    """Copy the live compounder_data.json into the archive folder, named
    by its own generated_at timestamp - called right before a rebuild
    REPLACES the live data, so every previous dataset stays viewable on
    the Research page's snapshot picker. Idempotent: re-archiving the
    same generated_at overwrites the same file rather than duplicating.
    Returns the archive path, or None if there was nothing to archive."""
    current_path = os.path.join(_cp_data_dir(), "compounder_data.json")
    if not os.path.exists(current_path):
        return None
    try:
        with open(current_path) as f:
            data = json.load(f)
    except (OSError, ValueError):
        return None
    gen = str(data.get("generated_at") or "unknown")
    safe = "".join(c if (c.isalnum() or c in "T-") else "-" for c in gen)[:19] or "unknown"
    dest = os.path.join(_cp_archive_dir(), f"compounder_data_{safe}.json")
    try:
        with open(dest, "w") as f:
            json.dump(data, f, indent=1)
    except OSError:
        return None
    return dest


def list_archived_snapshots():
    """Archived datasets, newest first: [{'path', 'generated_at',
    'label', 'n_tickers'}]. Labels are human-readable ('11 Aug 2026,
    03:58 UTC - 3 companies') for the Research page's snapshot picker."""
    from datetime import datetime as _dt, timezone as _tz
    out = []
    try:
        files = sorted(os.listdir(_cp_archive_dir()), reverse=True)
    except OSError:
        return out
    for fn in files:
        if not (fn.startswith("compounder_data_") and fn.endswith(".json")):
            continue
        path = os.path.join(_cp_archive_dir(), fn)
        try:
            with open(path) as f:
                data = json.load(f)
        except (OSError, ValueError):
            continue
        gen = data.get("generated_at")
        try:
            dt = _dt.fromisoformat(gen)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=_tz.utc)
            when = dt.astimezone(_tz.utc).strftime("%d %b %Y, %H:%M UTC")
        except (TypeError, ValueError):
            when = "unknown date"
        n = len(data.get("tickers", {}))
        out.append({
            "path": path, "generated_at": gen,
            "label": f"{when} - {n} companies", "n_tickers": n,
        })
    out.sort(key=lambda s: s.get("generated_at") or "", reverse=True)
    return out


def load_snapshot(path):
    """One archived dataset, or None if unreadable."""
    try:
        with open(path) as f:
            return json.load(f)
    except (OSError, ValueError):
        return None


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_PATH
    data = build(path)
    with open(OUT_PATH, "w") as f:
        json.dump(data, f, indent=1)
    n_tickers = len(data["tickers"])
    n_metrics = sum(len(s["metrics"]) for s in data["sections"].values())
    print(f"Wrote {OUT_PATH}: {n_tickers} tickers, {n_metrics} metrics across {len(data['sections'])} sections.")
